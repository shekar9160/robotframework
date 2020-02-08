import pandas as pd

#File Path
path = "/home/vidyayug-d7/Hidden-all.xlsx"

#To read Excel file
'''
df = pd.read_excel(path)
columns = df.head()
print(columns)
'''

#Example 2
'''
cols = [1, 2, 3]

df = pd.read_excel(path, sheet_names=['Sheet10','All'], na_values="Missing", skiprows=0,  usecols=cols)
columns = df.head()
print(columns)
'''
#Example 3
'''
df = pd.read_excel(path, sheet_name=['Sheet10','All'], na_values="Missing")
columns = df
print(columns)

'''
'''
#Example 4 (Read All Sheets)
all_sheets_df = pd.read_excel(path, sheet_name=None)
print(all_sheets_df)
'''
#Example 5 (Read All Sheets)
'''
df = pd.read_excel(path, sheet_name='All', header=0 ,dtype={'Forum_UUID':str,'Forum_Name':str, 'Hidden_content':str, 'Reason to hide content':str})
#To see excel file information
print(df.info())
print(df)

'''
#Example 5 (Creating Excel with data if not exist || just append values)with column header & File will be created automatically
'''
df = pd.DataFrame({'Names':['Andreas', 'George', 'Steve', 'Sarah', 'Joanna', 'Hanna'], 'Age':[21, 22, 20, 19, 18, 23]})
df.to_excel('NamesAndAges.xlsx')

'''
#Example 5 (Read All Sheets)
#'''
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
 
#df = pd.DataFrame({'Names':['Andreas', 'George', 'Steve', 'Sarah', 'Joanna', 'Hanna'], 'Age':[21, 22, 20, 19, 18, 23]})
df = pd.DataFrame({'Names':['Shekar', 'Ranju'], 'Age':[ 28, 33]})
#df.to_excel('NamesAndAges.xlsx')   
#append_df_to_excel('NamesAndAges.xlsx', df, sheet_name='Sheet1',truncate_sheet=True)
append_df_to_excel('NamesAndAges.xlsx', df, sheet_name='Sheet2', startrow=None, truncate_sheet=False, index=False)

