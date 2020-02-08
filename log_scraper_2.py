from bs4 import BeautifulSoup
from selenium import webdriver
import os
import pandas as pd

#############################
###Define system values
#############################
sys_str = "Comes from System"
empty_str = " "

##############################
##  Excel File settings    ###
##############################
filename = "ImageView Rohrbach RTM_Template for Mike A.xlsx"
sheet = "2.1 Requirements Trace Matrix"



#https://medium.com/ymedialabs-innovation/web-scraping-using-beautiful-soup-and-selenium-for-dynamic-page-2f8ad15efe25
#To read log file
#filepath = "file:///home/vidyayug-d7/Downloads/log-copy.html"
filepath = "file:///home/vidyayug-d7/Downloads/log6.html"

driver = webdriver.Firefox()
driver.get(filepath)
page_source = driver.page_source

expand_elements = driver.find_elements_by_xpath('//div[@class="element-header closed"]')
for element in expand_elements:
    element.click()
    
page_source = driver.page_source
#print("--------------------------------------------------------------------------")
#print(page_source.encode('utf-8'))
soup = BeautifulSoup(page_source, "lxml")
test_suites = soup.find_all('div', attrs={"class":["suite"]})
#test_cases = soup.find_all('div', class_='test')
#print(test_suites)
print("There are {} Test cases.".format(len(test_suites)))


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

#Extract information
for suit in test_suites[1:]:
    print("--------------------------------------------------------------------------")
    test_case=""
    d_id = ""
    d_dir = ""
    result = ""
    status = ""
    try:
        if suit.find('span', class_='name'):
            test_case = suit.find('span', class_='name').text
        if len(suit.find_all('td', class_='doc')) > 0:
            raw_id = suit.find_all('td', class_='doc')[1]
            d_id = raw_id.find('p').text
            #print(d_id)
            str_split = d_id.split(":")
            d_did = str_split[0]
            d_dir = str_split[1]
        
        dcase = suit.find_all("th", string="Tags:")
        print("Tags :", dcase[0].find_next('td').text)
        tags = suit.find_all("th", string="Status:")      
        #print("Status :", tags[1].find_next('span', class_='label').text)
 
        if dcase[0].find_next('td').contents[0]:
            result = dcase[0].find_next('td').text            
            #if result.find("DE456") == 1:
            if result.find("DE3576") == -1:
                print("No")
            else:
                print("Yes")           
        if tags[1].find_next('span', class_='label'):                
            status = tags[1].find_next('span', class_='label').text
        
        print("Test case: {}".format(test_case))
        print("Design Input ID: {}".format(d_did))
        print("Design Input Requirement: {}".format(d_dir))
        print("Defect case: {}".format(result))
        print("Status: {}".format(status))
        data = [[d_did, d_dir, sys_str, test_case, sys_str, sys_str, sys_str, status, result]]
        df = pd.DataFrame(data)

        #Now Store values to Excel
        append_df_to_excel("ImageView Rohrbach RTM_Template for Mike A.xlsx", df, sheet_name="2.1 Requirements Trace Matrix", startrow=None, truncate_sheet=False, index=False, header=None)
    except Exception as e:
        print(e)
        pass

'''
#Now Store values to Excel

d_did = d_dir = test_plan_id = test_Case = test_cysle = test_cycle_results = release_cycle = status = result = ""

#order : d_did, d_dir, test_plan_id, test_Case, test_cysle, test_cycle_results, release_cycle, status, result
#yes, yes, no, yes, no, no, no, yes, yes
data = [[d_did, d_dir, '', test_case, '', '', '', status, result]]
df = pd.DataFrame(data)

#Now Store values to Excel
append_df_to_excel("sample.xlsx", df, sheet_name="Sheet1", startrow=None, truncate_sheet=False, index=False, header=None)
''' 
driver.quit()
#driver.close()

